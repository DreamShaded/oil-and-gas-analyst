from __future__ import annotations

import asyncio
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
from playwright.async_api import async_playwright

from src.data.cache import FileCache
from src.utils.logging import get_logger

log = get_logger("data.minfin_oil")

STATS_PAGE_URL = "https://minfin.gov.ru/ru/statistics/fedbud/oil/"
USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)
XLSX_TTL = timedelta(days=14)


async def _discover_xlsx() -> str | None:
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            ctx = await browser.new_context(user_agent=USER_AGENT)
            page = await ctx.new_page()
            await page.goto(STATS_PAGE_URL, wait_until="networkidle", timeout=30_000)
            await page.wait_for_timeout(1500)
            links: list[str] = await page.eval_on_selector_all(
                "a[href$='.xlsx']",
                "els => Array.from(new Set(els.map(e => e.href)))",
            )
        finally:
            await browser.close()
    relevant = [u for u in links if "neftegaz" in u.lower() or "oil" in u.lower() or "Svedeniya" in u]
    return (relevant or links or [None])[0]


async def _download(url: str, target: Path) -> int:
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        try:
            ctx = await browser.new_context(user_agent=USER_AGENT)
            await ctx.new_page()
            resp = await ctx.request.get(url, timeout=60_000)
            if resp.status != 200:
                raise RuntimeError(f"Minfin xlsx: HTTP {resp.status}")
            data = await resp.body()
        finally:
            await browser.close()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(data)
    return len(data)


def parse_xlsx(xlsx_path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next((i for i, r in enumerate(rows) if r and r[0] == "Показатели"), None)
    if header_idx is None:
        raise RuntimeError("Minfin xlsx: не нашёл строку 'Показатели'")
    months = [c for c in rows[header_idx][1:] if c is not None]
    records: list[dict] = []
    for r in rows[header_idx + 2:]:
        if not r or r[0] is None or str(r[0]).strip() in {"", "в том числе за счет:"}:
            continue
        metric = str(r[0]).strip()
        for col_idx, month_ts in enumerate(months, start=1):
            value = r[col_idx] if col_idx < len(r) else None
            if value in (None, "-", ""):
                continue
            try:
                num = float(str(value).replace(",", "."))
            except (TypeError, ValueError):
                continue
            period = month_ts.date() if hasattr(month_ts, "date") else month_ts
            if not isinstance(period, date):
                continue
            records.append({"period": period, "metric": metric, "value": num})
    return pd.DataFrame(records)


def fetch_minfin_oil(out_dir: Path, *, cache_root: Path | None = None) -> pd.DataFrame:
    cache = FileCache(cache_root or Path("data/cache"))
    out_dir.mkdir(parents=True, exist_ok=True)

    url = asyncio.run(_discover_xlsx())
    if not url:
        log.warning("minfin.no_xlsx_link")
        return pd.DataFrame()

    filename = url.rsplit("/", 1)[-1]
    target = out_dir / filename
    cache_key = f"minfin/{filename}"
    if not (target.exists() and cache.is_fresh(cache_key, XLSX_TTL)):
        size = asyncio.run(_download(url, target))
        cache.write_bytes(cache_key, b"ok")
        log.info("minfin.downloaded", file=filename, bytes=size)
    else:
        log.info("minfin.skip_fresh", file=filename)

    df = parse_xlsx(target)
    df.to_csv(out_dir / "oil_revenue.csv", index=False)
    log.info("minfin.parsed", rows=len(df), metrics=df["metric"].nunique() if not df.empty else 0)
    return df
